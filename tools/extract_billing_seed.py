import json
import re
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "Sai Nivas expendeature.xlsx"
OUTPUT = ROOT / "billing-seed.json"

MONTHS = {
    "jan": 1,
    "january": 1,
    "feb": 2,
    "february": 2,
    "mar": 3,
    "march": 3,
    "apr": 4,
    "april": 4,
    "may": 5,
    "jun": 6,
    "june": 6,
    "jul": 7,
    "july": 7,
    "aug": 8,
    "august": 8,
    "sep": 9,
    "sept": 9,
    "september": 9,
    "oct": 10,
    "october": 10,
    "nov": 11,
    "november": 11,
    "dec": 12,
    "december": 12,
}


def clean_value(value):
    if value is None:
        return None
    if isinstance(value, float):
        return round(value, 6)
    return value


def month_key(sheet_name, order):
    normalized = re.sub(r"[\s_\-()]+", " ", sheet_name.strip().lower())
    match = re.search(r"(jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|jul(?:y)?|aug(?:ust)?|sep(?:t|tember)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?)\s*(\d{4})", normalized)
    if match:
      month_num = MONTHS[match.group(1)]
      return f"{int(match.group(2)):04d}-{month_num:02d}"
    return f"sheet-{order:03d}"


def parse_expense_sections(ws):
    sections = []
    for col in range(12, ws.max_column + 1):
        title = ws.cell(1, col).value
        if isinstance(title, str) and "expend" in title.lower():
            items = []
            for row in range(2, ws.max_row + 1):
                label = ws.cell(row, col).value
                amount = ws.cell(row, col + 1).value if col + 1 <= ws.max_column else None
                status = ws.cell(row, col + 2).value if col + 2 <= ws.max_column else None
                if label is None and amount is None and status is None:
                    continue
                items.append(
                    {
                        "label": clean_value(label),
                        "amount": clean_value(amount),
                        "status": clean_value(status),
                    }
                )
            sections.append({"title": title, "items": items})
    return sections


def parse_month_sheet(ws, order):
    if ws["A1"].value != "Flat Numbers":
        return None

    flats = []
    for row in range(2, 21):
        flat_number = ws.cell(row, 1).value
        if flat_number is None:
            continue

        topay = ws.cell(row, 9).value
        receive_amount = ws.cell(row, 10).value
        payment_status = ws.cell(row, 11).value

        if isinstance(receive_amount, str) and payment_status is None:
            payment_status = receive_amount
            receive_amount = None

        flats.append(
            {
                "flatNumber": str(flat_number),
                "previousReading": clean_value(ws.cell(row, 2).value),
                "currentReading": clean_value(ws.cell(row, 3).value),
                "totalUnits": clean_value(ws.cell(row, 4).value),
                "waterBill": clean_value(ws.cell(row, 5).value),
                "commonMaintenance": clean_value(ws.cell(row, 6).value),
                "totalMaintenance": clean_value(ws.cell(row, 7).value),
                "garbageAmount": clean_value(ws.cell(row, 8).value),
                "toPay": clean_value(topay),
                "receiveAmount": clean_value(receive_amount),
                "paymentStatus": clean_value(payment_status),
            }
        )

    expenses = parse_expense_sections(ws)
    totals_row = 22
    rate_row = 23
    tanker_row = 24
    tanker_text = ws.cell(tanker_row, 3).value
    tanker_count = None
    tanker_cost = None
    if isinstance(tanker_text, str):
        count_match = re.search(r"Tankers\((\d+)\)", tanker_text, re.IGNORECASE)
        cost_match = re.search(r"Cost:\s*(\d+)", tanker_text, re.IGNORECASE)
        if count_match:
            tanker_count = int(count_match.group(1))
        if cost_match:
            tanker_cost = int(cost_match.group(1))

    return {
        "id": month_key(ws.title, order),
        "label": ws.title,
        "sheetName": ws.title,
        "carryForwardLabel": clean_value(ws["L1"].value),
        "carryForwardValue": clean_value(ws["M1"].value),
        "summary": {
            "garbageTotal": clean_value(ws.cell(21, 8).value),
            "tableTotalToPay": clean_value(ws.cell(21, 9).value),
            "tableTotalReceived": clean_value(ws.cell(21, 10).value),
            "totalUsageUnits": clean_value(ws.cell(totals_row, 4).value),
            "totalWaterBill": clean_value(ws.cell(totals_row, 5).value),
            "perUnitCost": clean_value(ws.cell(rate_row, 4).value),
            "declaredWaterTotal": clean_value(ws.cell(tanker_row, 2).value),
            "tankerCount": tanker_count,
            "costPerTanker": tanker_cost,
        },
        "expenseSections": expenses,
        "flats": flats,
    }


def build_seed_from_workbook(workbook_path=WORKBOOK, output_path=OUTPUT):
    workbook = load_workbook(workbook_path, data_only=True)
    months = []

    for order, name in enumerate(workbook.sheetnames, start=1):
        parsed = parse_month_sheet(workbook[name], order)
        if parsed:
            months.append(parsed)

    seed = {
        "fixedFlats": [flat["flatNumber"] for flat in months[-1]["flats"]],
        "months": months,
        "latestMonthId": months[-1]["id"] if months else None,
    }

    output_path.write_text(json.dumps(seed, indent=2), encoding="utf-8")
    return seed


def build_seed():
    build_seed_from_workbook(WORKBOOK, OUTPUT)
    print(f"Wrote {OUTPUT}")


if __name__ == "__main__":
    build_seed()
