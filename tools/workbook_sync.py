import json
import re
from copy import copy
from pathlib import Path

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "Sai Nivas expendeature.xlsx"
SEED_FILE = ROOT / "billing-seed.json"

HEADERS = [
    "Flat Numbers",
    "Previous Meter Reading",
    "Current Meter Reading",
    "Total Units",
    "Water Bill",
    "Common Maintanence",
    "Total Maintenance",
    "Garbage Amount",
    "Topay ",
    "Receive Amount",
    "Payment Status",
]


def month_label(month_id):
    year, month = [int(part) for part in month_id.split("-")]
    names = [
        "January",
        "February",
        "March",
        "April",
        "May",
        "June",
        "July",
        "August",
        "September",
        "October",
        "November",
        "December",
    ]
    return f"{names[month - 1]} {year}"


def normalize_sheet_name(name):
    return re.sub(r"[\s_\-()]+", " ", str(name).strip().lower())


def month_key_from_name(name):
    normalized = normalize_sheet_name(name)
    match = re.search(
        r"(jan(?:uary)?|feb(?:ruary)?|mar(?:ch)?|apr(?:il)?|may|jun(?:e)?|jul(?:y)?|aug(?:ust)?|sep(?:t|tember)?|oct(?:ober)?|nov(?:ember)?|dec(?:ember)?)\s*(\d{4})",
        normalized,
    )
    if not match:
        return None

    month_map = {
        "jan": 1,
        "january": 1,
        "feb": 2,
        "february": 2,
        "mar": 3,
        "march": 3,
        "apr": 4,
        "april": 4,
        "may": 5,
        "jun": 6,
        "june": 6,
        "jul": 7,
        "july": 7,
        "aug": 8,
        "august": 8,
        "sep": 9,
        "sept": 9,
        "september": 9,
        "oct": 10,
        "october": 10,
        "nov": 11,
        "november": 11,
        "dec": 12,
        "december": 12,
    }
    return f"{int(match.group(2)):04d}-{month_map[match.group(1)]:02d}"


def recent_month_sheets(workbook):
    items = []
    for sheet_name in workbook.sheetnames:
        key = month_key_from_name(sheet_name)
        if key:
            items.append((key, sheet_name))
    return sorted(items, key=lambda item: item[0])


def clear_sheet(ws):
    for merged in list(ws.merged_cells.ranges):
        ws.unmerge_cells(str(merged))
    for row in ws.iter_rows():
        for cell in row:
            cell.value = None


def copy_sheet_dimensions(source, target):
    for key, value in source.column_dimensions.items():
        target.column_dimensions[key] = copy(value)
    for key, value in source.row_dimensions.items():
        target.row_dimensions[key] = copy(value)
    for merged in list(source.merged_cells.ranges):
        target.merge_cells(str(merged))


def template_sheet(workbook):
    month_sheets = recent_month_sheets(workbook)
    if not month_sheets:
        return None
    return workbook[month_sheets[-1][1]]


def set_if_present(ws, cell_ref, value):
    if value is not None and value != "":
        ws[cell_ref] = value


def water_bill_note(month):
    return f"WATER Tankers({int(month['settings'].get('tankerCount', 0) or 0)}) One Tanker Cost: {int(month['settings'].get('costPerTanker', 0) or 0)}"


def compute_month_values(month):
    flats = month["flats"]
    tanker_count = float(month["settings"].get("tankerCount", 0) or 0)
    cost_per_tanker = float(month["settings"].get("costPerTanker", 0) or 0)
    total_water_cost = tanker_count * cost_per_tanker

    def is_shop(name):
        return str(name).lower().startswith("shop")

    def is_common(name):
        return str(name).lower() == "common"

    def to_num(value):
        if value in (None, "", "NA"):
            return None
        try:
            return float(value)
        except (TypeError, ValueError):
            return None

    rows = []
    for flat in flats:
        prev = to_num(flat.get("previousReading"))
        curr = to_num(flat.get("currentReading"))
        units = 0 if prev is None or curr is None else max(round((curr - prev) * 1000), 0)
        rows.append(
            {
                **flat,
                "units": units,
            }
        )

    total_usage_units = sum(row["units"] for row in rows)
    per_unit_cost = (total_water_cost / total_usage_units) if total_usage_units else 0
    common_row = next((row for row in rows if is_common(row["flatNumber"])), None)
    common_water_bill = round((common_row["units"] if common_row else 0) * per_unit_cost)
    residential = [row for row in rows if not is_shop(row["flatNumber"]) and not is_common(row["flatNumber"])]
    share_base = (common_water_bill // len(residential)) if residential else 0
    share_remainder = (common_water_bill % len(residential)) if residential else 0

    for index, row in enumerate(residential):
        row["commonShare"] = share_base + (1 if index < share_remainder else 0)

    for row in rows:
        if is_shop(row["flatNumber"]):
            row["baseWaterBill"] = 0
            row["commonShare"] = 0
            row["garbageAmount"] = 0
        elif is_common(row["flatNumber"]):
            row["baseWaterBill"] = 0
            row["commonShare"] = 0
            row["garbageAmount"] = 0
            row["commonMaintenance"] = 0
        else:
            row["baseWaterBill"] = round(row["units"] * per_unit_cost)
            row["commonShare"] = row.get("commonShare", 0)

        row["commonMaintenance"] = float(row.get("commonMaintenance") or 0)
        row["receivedAmount"] = float(row.get("receivedAmount") or 0)
        row["totalMaintenance"] = row["baseWaterBill"] + row["commonShare"] + row["commonMaintenance"]
        row["totalDue"] = 0 if is_common(row["flatNumber"]) else row["totalMaintenance"] + float(row.get("garbageAmount") or 0)
        row["paymentStatus"] = row.get("paymentStatus") or ""

    return {
        "rows": rows,
        "total_usage_units": total_usage_units,
        "per_unit_cost": per_unit_cost,
        "total_water_cost": total_water_cost,
        "common_water_bill": common_water_bill,
        "garbage_total": sum(float(row.get("garbageAmount") or 0) for row in rows),
        "total_due": sum(row["totalDue"] for row in rows),
        "total_received": sum(float(row.get("receivedAmount") or 0) for row in rows),
    }


def write_month_sheet(workbook, month):
    existing = workbook[month["label"]] if month["label"] in workbook.sheetnames else None
    if existing is None:
        template = template_sheet(workbook)
        ws = workbook.create_sheet(month["label"])
        if template is not None:
            copy_sheet_dimensions(template, ws)
    else:
        ws = existing

    clear_sheet(ws)
    month_values = compute_month_values(month)

    for index, header in enumerate(HEADERS, start=1):
        ws.cell(1, index).value = header

    carry_label = month.get("carryForwardLabel") or f"Bank balance as of {month_label(month['id'])}"
    ws["L1"] = carry_label
    ws["M1"] = month.get("carryForwardValue", 0)

    for row_index, row in enumerate(month_values["rows"], start=2):
        flat_number = row["flatNumber"]
        shop = str(flat_number).lower().startswith("shop")
        common = str(flat_number).lower() == "common"

        ws.cell(row_index, 1).value = flat_number
        ws.cell(row_index, 2).value = "NA" if shop else row.get("previousReading")
        ws.cell(row_index, 3).value = "NA" if shop else row.get("currentReading")
        ws.cell(row_index, 4).value = "NA" if shop else row["units"]
        ws.cell(row_index, 5).value = "NA" if shop else (0 if common else row["baseWaterBill"])
        ws.cell(row_index, 6).value = 300 if shop else row["commonMaintenance"]
        ws.cell(row_index, 7).value = 300 if shop else (None if common else row["totalMaintenance"])
        ws.cell(row_index, 8).value = 0 if shop or common else row["garbageAmount"]
        ws.cell(row_index, 9).value = 300 if shop else (0 if common else row["totalDue"])
        ws.cell(row_index, 10).value = None if common else (row["receivedAmount"] or None)
        ws.cell(row_index, 11).value = None if common else row["paymentStatus"]

    totals_row = 21
    ws.cell(totals_row, 8).value = month_values["garbage_total"]
    ws.cell(totals_row, 9).value = month_values["total_due"]
    ws.cell(totals_row, 10).value = month_values["total_received"]

    ws.cell(22, 2).value = "Total Usage Units"
    ws.cell(22, 4).value = month_values["total_usage_units"]
    ws.cell(22, 5).value = round(month_values["total_water_cost"])

    ws.cell(23, 2).value = "Per Lt cost"
    ws.cell(23, 3).value = "Number of tankers / Number of litters"
    ws.cell(23, 4).value = month_values["per_unit_cost"]

    ws.cell(24, 1).value = "Total Water Bill"
    ws.cell(24, 2).value = round(month_values["total_water_cost"])
    ws.cell(24, 3).value = water_bill_note(month)

    expenses = month.get("expenses", [])
    if expenses:
        ws.cell(1, 15).value = month.get("expenseTitle") or f"{month_label(month['id'])} Expenditures"
        ws.cell(1, 16).value = "Amount"
        ws.cell(1, 17).value = "Status"
        for row_index, item in enumerate(expenses, start=2):
            ws.cell(row_index, 15).value = item.get("label") or None
            ws.cell(row_index, 16).value = item.get("amount") or None
            ws.cell(row_index, 17).value = item.get("status") or None


def build_seed_from_state(state):
    seed_months = []
    for month in state["months"]:
      seed_months.append(
          {
              "id": month["id"],
              "label": month["label"],
              "sheetName": month["label"],
              "carryForwardLabel": month.get("carryForwardLabel"),
              "carryForwardValue": month.get("carryForwardValue"),
              "summary": {},
              "expenseSections": [
                  {
                      "title": month.get("expenseTitle") or f"{month_label(month['id'])} Expenditures",
                      "items": month.get("expenses", []),
                  }
              ],
              "flats": month["flats"],
          }
      )

    return {
        "fixedFlats": state.get("fixedFlats", []),
        "months": seed_months,
        "latestMonthId": state.get("selectedMonthId"),
    }


def sync_state_to_workbook(state):
    workbook = load_workbook(WORKBOOK)
    month_names = {month_key_from_name(name): name for name in workbook.sheetnames if month_key_from_name(name)}

    for month in state["months"]:
        if month["id"] in month_names and month_names[month["id"]] != month["label"]:
            ws = workbook[month_names[month["id"]]]
            ws.title = month["label"]
        write_month_sheet(workbook, month)

    workbook.save(WORKBOOK)
    SEED_FILE.write_text(json.dumps(build_seed_from_state(state), indent=2), encoding="utf-8")
